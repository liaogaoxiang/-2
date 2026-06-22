import base64
import gzip
import json
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError as exc:
    raise SystemExit("openpyxl is required. Use the bundled Codex Python runtime shown in README.md.") from exc


REPORT_PATH = Path(sys.argv[1] if len(sys.argv) > 1 else "data/report-data.js")
ARCH_PATH = Path(sys.argv[2] if len(sys.argv) > 2 else "/Users/gaoxiangliao/Desktop/精品架构映射.xlsx")
OUTPUT_PATH = Path("data/management-data.js")
DEPARTMENT_OUTPUTS = {
    "学习顾问部": Path("data/management-advisor-data.js"),
    "学习规划部": Path("data/management-planning-data.js"),
}
ROLE_OUTPUTS = {
    ("学习顾问部", "manager"): Path("data/management-advisor-manager-data.js"),
    ("学习顾问部", "n2"): Path("data/management-advisor-n2-data.js"),
    ("学习规划部", "manager"): Path("data/management-planning-manager-data.js"),
    ("学习规划部", "n2"): Path("data/management-planning-n2-data.js"),
}
TARGET_GMV = 60000
SPRINT_GMV = 50000
ROLE_PLACEHOLDERS = {"未匹配架构", "未归属", "其他", ""}

N2_LADDER = [
    {
        "threshold": 3,
        "title": "3组档",
        "prizes": ["华为手环10", "防晒衣", "电饼铛", "奥克斯蒸烤炸三合一", "美的破壁机"],
    },
    {
        "threshold": 6,
        "title": "6组档",
        "prizes": ["联想平板", "AirPods", "华为AI眼镜", "米家打印机Pro", "小米冷风扇"],
    },
    {
        "threshold": 9,
        "title": "9组档",
        "prizes": ["华为Pad", "华为手机70s", "小米AI智能家庭屏幕", "惠普打印机", "美的95升冰吧"],
    },
    {
        "threshold": 12,
        "title": "12组档",
        "prizes": ["九号电动滑板车", "小米55英寸电视", "大疆Action 4", "小米双开门冰箱", "科沃斯扫地机"],
    },
]

MANAGER_LADDER = [
    {
        "threshold": 8,
        "title": "8组档",
        "prizes": ["华为手环10", "防晒衣", "电饼铛", "奥克斯蒸烤炸三合一", "美的破壁机"],
    },
    {
        "threshold": 12,
        "title": "12组档",
        "prizes": ["联想平板", "AirPods", "华为AI眼镜", "米家打印机Pro", "小米冷风扇"],
    },
    {
        "threshold": 16,
        "title": "16组档",
        "prizes": ["华为Pad", "华为手机70s", "小米AI智能家庭屏幕", "惠普打印机", "美的95升冰吧"],
    },
    {
        "threshold": 20,
        "title": "20组档",
        "prizes": ["九号电动滑板车", "小米55英寸电视", "大疆Action 4", "小米双开门冰箱", "科沃斯扫地机"],
    },
    {
        "threshold": 22,
        "title": "22组档",
        "prizes": ["苹果手表11最新款", "大疆pocket4", "ipad最新款128G"],
    },
]


def load_report(path):
    text = path.read_text(encoding="utf-8")
    match = re.search(r'atob\("([^"]+)"\)', text)
    if not match:
        raise SystemExit(f"Could not find compressed report payload in {path}")
    return json.loads(gzip.decompress(base64.b64decode(match.group(1))).decode("utf-8"))


def clean(value, fallback=""):
    text = str(value or "").strip()
    if not text or text in {"-", "null", "None"}:
        return fallback
    return text


def load_architecture(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    required = {"n1", "n2", "manager", "department_1", "department_2", "department_3"}
    selected = None
    selected_index = None
    for ws in wb.worksheets:
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        index = {name: pos for pos, name in enumerate(headers)}
        if required.issubset(index):
            selected = ws
            selected_index = index
            break
    if selected is None:
        first_headers = [cell.value for cell in next(wb.active.iter_rows(min_row=1, max_row=1))]
        index = {name: pos for pos, name in enumerate(first_headers)}
    else:
        ws = selected
        index = selected_index
    missing = required - set(index)
    if missing:
        raise SystemExit(f"Architecture file is missing required columns: {', '.join(sorted(missing))}")

    mapping = defaultdict(lambda: {"n2": Counter(), "manager": Counter(), "module": Counter(), "members": 0})
    for row in ws.iter_rows(min_row=2, values_only=True):
        n1 = clean(row[index["n1"]])
        if not n1:
            continue
        n2 = clean(row[index["n2"]], "未归属")
        manager = clean(row[index["manager"]], "未归属")
        department = " ".join(clean(row[index[name]]) for name in ("department_1", "department_2", "department_3"))
        if "学习顾问部" in department:
            module = "学习顾问部"
        elif "学习规划部" in department:
            module = "学习规划部"
        else:
            module = "其他"
        mapping[n1]["n2"][n2] += 1
        mapping[n1]["manager"][manager] += 1
        mapping[n1]["module"][module] += 1
        mapping[n1]["members"] += 1
    return mapping


def top(counter, fallback):
    return counter.most_common(1)[0][0] if counter else fallback


def architecture_meta(mapping, name, fallback_module):
    row = mapping.get(name)
    if not row or not row["members"]:
        return {
            "n2": "未匹配架构",
            "manager": "未匹配架构",
            "module": fallback_module,
            "archMembers": 0,
            "mapped": False,
        }
    return {
        "n2": top(row["n2"], "未归属"),
        "manager": top(row["manager"], "未归属"),
        "module": top(row["module"], fallback_module),
        "archMembers": row["members"],
        "mapped": True,
    }


def ladder_status(count, ladder):
    reached = None
    for tier in ladder:
        if count >= tier["threshold"]:
            reached = tier
    next_tier = next((tier for tier in ladder if count < tier["threshold"]), None)
    return {
        "reached": reached["threshold"] if reached else 0,
        "reachedTitle": reached["title"] if reached else "未达档",
        "next": next_tier["threshold"] if next_tier else None,
        "nextTitle": next_tier["title"] if next_tier else "满档",
        "gap": max(0, next_tier["threshold"] - count) if next_tier else 0,
        "isMax": next_tier is None,
    }


def apply_upgrade(status, qualified, ladder):
    reward_threshold = status["reached"]
    reward_title = status["reachedTitle"]
    upgraded = False

    if qualified and status["reached"]:
        current_index = next(
            (index for index, tier in enumerate(ladder) if tier["threshold"] == status["reached"]),
            None,
        )
        if current_index is not None and current_index + 1 < len(ladder):
            reward_threshold = ladder[current_index + 1]["threshold"]
            reward_title = ladder[current_index + 1]["title"]
            upgraded = True

    return {
        **status,
        "upgradeQualified": qualified,
        "upgraded": upgraded,
        "rewardThreshold": reward_threshold,
        "rewardTitle": reward_title,
        "upgradedTitle": f"奖励升至{reward_title}" if upgraded else reward_title,
    }


def aggregate(leaders, group_key, ladder):
    rows = {}
    for leader in leaders.values():
        key = (leader["module"], leader[group_key])
        row = rows.setdefault(
            key,
            {
                "module": leader["module"],
                "name": leader[group_key],
                "totalGroups": 0,
                "achievedGroups": 0,
                "totalGmv": 0.0,
                "achievedGmv": 0.0,
                "missingGroups": 0,
                "achievedTeams": [],
                "sprintTeams": [],
                "idleTeams": [],
            },
        )
        row["totalGroups"] += 1
        row["totalGmv"] += leader["gross"]
        if not leader["mapped"]:
            row["missingGroups"] += 1
        if leader["achieved"]:
            row["achievedGroups"] += 1
            row["achievedGmv"] += leader["gross"]
            row["achievedTeams"].append(leader)
        elif leader["gross"] >= SPRINT_GMV:
            row["sprintTeams"].append(leader)
        else:
            row["idleTeams"].append(leader)

    result = []
    for row in rows.values():
        row["achievedTeams"].sort(key=lambda item: (-item["gross"], item["name"]))
        row["sprintTeams"].sort(key=lambda item: (TARGET_GMV - item["gross"], item["name"]))
        row["idleTeams"].sort(key=lambda item: (-item["gross"], item["name"]))
        row["achievementRate"] = round(row["achievedGroups"] / row["totalGroups"], 4) if row["totalGroups"] else 0
        row["averageGmv"] = round(row["totalGmv"] / row["totalGroups"], 2) if row["totalGroups"] else 0
        row["status"] = apply_upgrade(
            ladder_status(row["achievedGroups"], ladder),
            row["achievementRate"] >= 0.7 and row["averageGmv"] > 70000,
            ladder,
        )
        row["totalGmv"] = round(row["totalGmv"], 2)
        row["achievedGmv"] = round(row["achievedGmv"], 2)
        result.append(row)

    result.sort(
        key=lambda item: (
            -item["achievedGroups"],
            item["status"]["gap"],
            -item["achievementRate"],
            -item["averageGmv"],
            item["module"],
            item["name"],
        )
    )
    return result


def team_summary(team):
    return {
        "name": team["name"],
        "module": team["module"],
        "n2": team["n2"],
        "manager": team["manager"],
        "gross": round(team["gross"], 2),
        "net": round(team["net"], 2),
        "orders": team["orders"],
        "gap": round(max(0, TARGET_GMV - team["gross"]), 2),
        "achieved": team["achieved"],
    }


def compact_row(row):
    return {
        **{key: row[key] for key in (
            "module",
            "name",
            "totalGroups",
            "achievedGroups",
            "totalGmv",
            "achievedGmv",
            "achievementRate",
            "averageGmv",
            "missingGroups",
            "status",
        )},
        "achievedTeams": [team_summary(team) for team in row["achievedTeams"]],
        "sprintTeams": [team_summary(team) for team in row["sprintTeams"][:8]],
        "idleTeams": [team_summary(team) for team in row["idleTeams"][:5]],
    }


def build_payload():
    report = load_report(REPORT_PATH)
    architecture = load_architecture(ARCH_PATH)
    leaders = {}

    for arena in report["arenas"]:
        for member in arena["members"]:
            name = member["name"]
            leader = leaders.setdefault(
                name,
                {
                    "name": name,
                    "gross": 0.0,
                    "net": 0.0,
                    "orders": 0,
                    "fallbackModule": arena["module"],
                },
            )
            leader["gross"] += member.get("gross") or 0
            leader["net"] += member.get("net") or 0
            leader["orders"] += member.get("orders") or 0

    for leader in leaders.values():
        leader.update(architecture_meta(architecture, leader["name"], leader["fallbackModule"]))
        leader["achieved"] = leader["gross"] >= TARGET_GMV

    raw_n2_rows = aggregate(leaders, "n2", N2_LADDER)
    manager_rows = [compact_row(row) for row in aggregate(leaders, "manager", MANAGER_LADDER)]
    manager_names = {row["name"] for row in manager_rows if row["name"] not in ROLE_PLACEHOLDERS}
    n2_conflict_rows = [
        compact_row(row)
        for row in raw_n2_rows
        if row["name"] in manager_names
    ]
    n2_rows = [
        compact_row(row)
        for row in raw_n2_rows
        if row["name"] not in manager_names
    ]
    achieved_teams = sorted((team_summary(row) for row in leaders.values() if row["achieved"]), key=lambda item: (-item["gross"], item["name"]))
    sprint_teams = sorted((team_summary(row) for row in leaders.values() if not row["achieved"] and row["gross"] >= SPRINT_GMV), key=lambda item: (item["gap"], item["name"]))
    missing_teams = sorted(
        (team_summary(row) for row in leaders.values() if not row["mapped"]),
        key=lambda item: item["name"],
    )

    return {
        "meta": {
            "title": "管理层荣耀补给战报",
            "module": "全部",
            "moduleTitle": "管理层",
            "sourceReport": str(REPORT_PATH),
            "architectureFile": ARCH_PATH.name,
            "reportDateRange": report["meta"]["dateRange"],
            "reportGeneratedAt": report["meta"].get("generatedAt"),
            "generatedAt": datetime.now(timezone.utc).isoformat(),
            "targetGmv": TARGET_GMV,
            "sprintGmv": SPRINT_GMV,
            "rule": "小组GMV达到6万即计入达标，不核验PK获胜；若同时为大组长和经理，按经理规则发奖",
            "n2ManagerDeduped": [row["name"] for row in n2_conflict_rows],
        },
        "ladders": {
            "n2": N2_LADDER,
            "manager": MANAGER_LADDER,
        },
        "summary": {
            "teams": len(leaders),
            "achievedTeams": len(achieved_teams),
            "sprintTeams": len(sprint_teams),
            "n2Unlocked": sum(1 for row in n2_rows if row["status"]["reached"] > 0),
            "managerUnlocked": sum(1 for row in manager_rows if row["status"]["reached"] > 0),
            "missingTeams": len(missing_teams),
            "totalGmv": round(sum(row["gross"] for row in leaders.values()), 2),
        },
        "n2Rows": n2_rows,
        "managerRows": manager_rows,
        "achievedTeams": achieved_teams,
        "sprintTeams": sprint_teams,
        "missingTeams": missing_teams,
    }


def filter_payload(payload, module):
    n2_rows = [row for row in payload["n2Rows"] if row["module"] == module]
    manager_rows = [row for row in payload["managerRows"] if row["module"] == module]
    achieved_teams = [team for team in payload["achievedTeams"] if team["module"] == module]
    sprint_teams = [team for team in payload["sprintTeams"] if team["module"] == module]
    missing_teams = [team for team in payload["missingTeams"] if team["module"] == module]
    module_payload = {
        **payload,
        "meta": {
            **payload["meta"],
            "title": f"{module}管理层荣耀补给战报",
            "module": module,
            "moduleTitle": module.replace("学习", ""),
        },
        "summary": {
            "teams": sum(row["totalGroups"] for row in n2_rows),
            "achievedTeams": len(achieved_teams),
            "sprintTeams": len(sprint_teams),
            "n2Unlocked": sum(1 for row in n2_rows if row["status"]["reached"] > 0),
            "managerUnlocked": sum(1 for row in manager_rows if row["status"]["reached"] > 0),
            "missingTeams": len(missing_teams),
            "totalGmv": round(sum(row["totalGmv"] for row in n2_rows), 2),
        },
        "n2Rows": n2_rows,
        "managerRows": manager_rows,
        "achievedTeams": achieved_teams,
        "sprintTeams": sprint_teams,
        "missingTeams": missing_teams,
    }
    return module_payload


def filter_role_payload(payload, module, role):
    module_payload = filter_payload(payload, module)
    row_key = "managerRows" if role == "manager" else "n2Rows"
    other_key = "n2Rows" if role == "manager" else "managerRows"
    rows = module_payload[row_key]
    role_title = "经理" if role == "manager" else "大组长"
    return {
        **module_payload,
        "meta": {
            **module_payload["meta"],
            "title": f"{module}{role_title}荣耀补给战报",
            "role": role,
            "roleTitle": role_title,
            "moduleRoleTitle": f"{module.replace('学习', '')}{role_title}",
        },
        "summary": {
            **module_payload["summary"],
            "teams": sum(row["totalGroups"] for row in rows),
            "achievedTeams": sum(row["achievedGroups"] for row in rows),
            "sprintTeams": sum(len(row["sprintTeams"]) for row in rows),
            "n2Unlocked": sum(1 for row in rows if row["status"]["reached"] > 0) if role == "n2" else 0,
            "managerUnlocked": sum(1 for row in rows if row["status"]["reached"] > 0) if role == "manager" else 0,
            "totalGmv": round(sum(row["totalGmv"] for row in rows), 2),
        },
        row_key: rows,
        other_key: [],
    }


def write_payload(path, payload):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        "window.PK_MANAGEMENT_DATA = "
        + json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
        + ";\n",
        encoding="utf-8",
    )


def main():
    payload = build_payload()
    write_payload(OUTPUT_PATH, payload)
    for module, output_path in DEPARTMENT_OUTPUTS.items():
        write_payload(output_path, filter_payload(payload, module))
    for (module, role), output_path in ROLE_OUTPUTS.items():
        write_payload(output_path, filter_role_payload(payload, module, role))
    print(
        f"Wrote {OUTPUT_PATH}: {payload['summary']['achievedTeams']} achieved teams, "
        f"{payload['summary']['n2Unlocked']} n2 unlocked, {payload['summary']['managerUnlocked']} managers unlocked"
    )


if __name__ == "__main__":
    main()
