import csv
import base64
import gzip
import json
from collections import defaultdict
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
import sys
import os

try:
    from openpyxl import load_workbook
except ImportError as exc:
    raise SystemExit("openpyxl is required. Use the bundled Codex Python runtime shown in README.md.") from exc


DATA_PATH = Path(sys.argv[1] if len(sys.argv) > 1 else os.getenv("PK_DATA_PATH", "input/0530.csv"))
GROUPING_PATH = Path(sys.argv[2] if len(sys.argv) > 2 else os.getenv("PK_GROUPING_PATH", "input/工作簿14.xlsx"))
OUTPUT_PATH = Path("data/report-data.js")
ALLOWED_MODULE_ORDER = ["学习顾问部", "学习规划部"]
ALLOWED_MODULES = set(ALLOWED_MODULE_ORDER)
ACTIVITY_NAMES = {
    "10216": "H业务线内部员工转介绍",
    "10227": "0228高中转介绍（适用在读&已毕业）",
    "10240": "高中体验课用户活动",
    "10275": "1对1-转介绍活动「1对1顾问」",
    "10390": "1对1转介绍活动【班主任】",
    "10404": "初/高中绑定关系过期补绑专用",
    "10418": "【H业务线】初中转介绍",
    "10435": "【2024志愿填报一对一】课程转介绍",
    "10496": "【H业务线】体验课用户转介绍初中",
    "10528": "1对1-顾问转介绍【适用非在读】",
    "10553": "1对1-班主任转介绍【非在读】",
    "10601": "【H业务线】语文项目部小学转介绍",
    "10628": "初中1对1-转介绍活动「1对1顾问」",
    "10717": "H-拼团活动",
    "10719": "清北校友会转介绍",
    "10380": "0701综合素养正价课转介绍",
    "10294": "高中转介绍途途小初专用",
    "10777": "KM综合素养线索转介绍",
    "10819": "初中1对1转介绍活动【班主任】",
    "10849": "跨学部转介绍【TT小初】",
    "10830": "【已停用别绑定】转介绍【青少小初】",
    "10896": "转介绍自媒体引流",
    "10890": "EM转介绍",
    "10928": "名师定制学用户转介绍",
    "10990": "2026高中毕业生转介绍",
}
ALLOWED_ACTIVITY_IDS = set(ACTIVITY_NAMES)

CSV_HEADERS = [
    "活动ID", "邀请人ID", "被邀请人ID", "邀请时间", "是否激励", "是否违规", "订单编号", "支付时间", "支付月份", "是否退款", "退款时间", "业绩归属人姓名", "业绩归属人人才类型", "业绩归属人部门_手工", "业绩归属人部门_系统", "年级_来自人员架构表", "中心_来自人员架构表", "经理_来自人员架构表", "大组长_来自人员架构表", "小组长_来自人员架构表", "课程一级部门名称", "课程二级部门名称", "课程三级部门名称", "课程一级科目名称", "课程二级科目名称", "课程三级科目名称", "课程类别名称", "年级", "科目", "学年", "学季", "学年学季", "主讲老师", "班级业务编号", "班级名称", "辅导班业务编码", "辅导老师", "原始订单编号", "原始支付时间", "最新订单编号", "最新用户编号", "绑定人ID", "绑定人姓名", "绑定人角色", "绑定人部门全路径", "跟进人ID", "跟进人姓名", "跟进人角色", "跟进人部门全路径", "绑定类型", "绑定类型名称", "模板ID", "海报URL", "风险因子", "用户意向", "标记", "业绩归属人部门_系统_用于净收完成度", "业绩归属人部门_手工_用于收款完成度", "绑定人部门", "收款金额", "退款金额", "净收金额"
]
FIELD_ALIASES = {
    "活动ID": ["活动id"],
    "被邀请人ID": ["用户id"],
    "邀请时间": ["邀请成功时间"],
    "是否激励": ["是否激励活动id"],
    "是否违规": ["is_violate"],
    "支付时间": ["转化时间"],
    "退款时间": ["最终退款时间"],
    "业绩归属人姓名": ["业绩归属人"],
    "经理_来自人员架构表": ["经理"],
    "大组长_来自人员架构表": ["大组长"],
    "小组长_来自人员架构表": ["小组长"],
    "课程二级部门名称": ["课程二级部门"],
    "课程三级部门名称": ["课程三级部门"],
    "课程一级科目名称": ["学科"],
    "科目": ["学科"],
    "主讲老师": ["主讲"],
    "班级业务编号": ["班级id"],
    "业绩归属人部门_系统": ["业绩归属人部门_系统"],
    "业绩归属人部门_手工": ["业绩归属人部门_手工"],
    "收款金额": ["收入", "订单金额"],
    "退款金额": ["退款"],
    "净收金额": ["净收"],
}


def empty_stat():
    return {"gross": 0.0, "refund": 0.0, "net": 0.0, "_orderUsers": set()}


def add_stat(stat, gross, refund, net, converted_user=""):
    if converted_user:
        stat["_orderUsers"].add(converted_user)
    stat["gross"] += gross
    stat["refund"] += refund
    stat["net"] += net


def clean_name(value, fallback="未归属"):
    text = str(value or "").strip()
    if not text or text in {"null", "-"}:
        return fallback
    return text


def clean_id(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def amount(value):
    text = str(value or "").strip().replace(",", "")
    if not text or text == "null":
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def round_stat(stat):
    orders = len(stat["_orderUsers"])
    return {
        "orders": orders,
        "gross": round(stat["gross"], 2),
        "refund": round(stat["refund"], 2),
        "net": round(stat["net"], 2),
        "refundRate": round(stat["refund"] / stat["gross"], 4) if stat["gross"] else 0,
        "avgNet": round(stat["net"] / orders, 2) if orders else 0,
    }


def normalize_csv_row(row):
    if len(row) == len(CSV_HEADERS):
        return row
    if len(row) > len(CSV_HEADERS):
        return row[:54] + [",".join(row[54 : len(row) - 7])] + row[len(row) - 7 :]
    return row + [""] * (len(CSV_HEADERS) - len(row))


def normalize_record(record):
    normalized = dict(record)
    for canonical, aliases in FIELD_ALIASES.items():
        if normalized.get(canonical) not in (None, ""):
            continue
        for alias in aliases:
            if record.get(alias) not in (None, ""):
                normalized[canonical] = record.get(alias)
                break

    if not normalized.get("支付月份") and normalized.get("支付时间"):
        normalized["支付月份"] = month_text(normalized.get("支付时间"))

    incentive_flag = str(normalized.get("是否激励") or "").strip().lower()
    if incentive_flag in {"1", "true", "yes"}:
        normalized["是否激励"] = "是"

    for header in CSV_HEADERS:
        normalized.setdefault(header, "")
    return normalized


def data_rows():
    if DATA_PATH.suffix.lower() == ".xlsx":
        wb = load_workbook(DATA_PATH, read_only=True, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        for values in ws.iter_rows(min_row=2, values_only=True):
            if any(value not in (None, "") for value in values):
                yield normalize_record(dict(zip(headers, values))), False
        return

    with DATA_PATH.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.reader(handle)
        header = next(reader)
        if header[: len(CSV_HEADERS)] != CSV_HEADERS[: len(header)]:
            raise SystemExit("CSV header does not match the expected export shape.")
        for raw in reader:
            if len(raw) == 1 and not raw[0]:
                continue
            repaired = len(raw) != len(CSV_HEADERS)
            yield normalize_record(dict(zip(CSV_HEADERS, normalize_csv_row(raw)))), repaired


def date_text(value):
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value or "")[:10]


def month_text(value):
    if isinstance(value, datetime):
        return value.strftime("%Y-%m")
    return str(value or "")[:7]


def row_module(row):
    department = clean_name(
        row.get("业绩归属人部门_手工") or row.get("业绩归属人部门_系统"),
        "",
    )
    return next((module for module in ALLOWED_MODULES if department.endswith(module)), "")


def parse_csv_stats():
    group_stats = defaultdict(empty_stat)
    group_converted_users = defaultdict(set)
    group_contributor_stats = defaultdict(lambda: defaultdict(empty_stat))
    group_meta = defaultdict(lambda: {"n2": Counter(), "manager": Counter()})
    contributor_rank_stats = defaultdict(empty_stat)
    contributor_modules = defaultdict(Counter)
    contributor_groups = defaultdict(Counter)
    daily_stats = defaultdict(empty_stat)
    subject_stats = defaultdict(empty_stat)
    activity_stats = defaultdict(empty_stat)
    total = empty_stat()
    months = set()
    repaired_rows = 0
    unassigned_rows = 0
    refunded_order_users = set()
    incentive_order_users = set()
    filtered_rows = 0
    activity_filtered_rows = 0
    converted_users = set()

    for row, repaired in data_rows():
        if repaired:
            repaired_rows += 1
        activity_id = clean_id(row.get("活动ID"))
        if activity_id not in ALLOWED_ACTIVITY_IDS:
            activity_filtered_rows += 1
            continue
        module = row_module(row)
        if not module:
            filtered_rows += 1
            continue
        gross = amount(row["收款金额"])
        refund = amount(row["退款金额"])
        net = amount(row["净收金额"])
        group = clean_name(row["小组长_来自人员架构表"])
        contributor = clean_name(row["业绩归属人姓名"], "未命名组员")
        converted_user = clean_name(row["被邀请人ID"], "")
        day = date_text(row["支付时间"])
        subject = clean_name(row["科目"] or row["课程一级科目名称"], "未知科目")

        add_stat(total, gross, refund, net, converted_user)
        add_stat(group_stats[group], gross, refund, net, converted_user)
        if converted_user:
            converted_users.add(converted_user)
            group_converted_users[group].add(converted_user)
        add_stat(group_contributor_stats[group][contributor], gross, refund, net, converted_user)
        add_stat(contributor_rank_stats[contributor], gross, refund, net, converted_user)
        contributor_modules[contributor][module] += 1
        contributor_groups[contributor][group] += 1
        group_meta[group]["n2"][clean_name(row["大组长_来自人员架构表"], "未知N2")] += 1
        group_meta[group]["manager"][clean_name(row["经理_来自人员架构表"], "未知管理者")] += 1
        add_stat(subject_stats[subject], gross, refund, net, converted_user)
        add_stat(activity_stats[activity_id], gross, refund, net, converted_user)
        if day:
            add_stat(daily_stats[day], gross, refund, net, converted_user)
        if row["支付月份"]:
            months.add(month_text(row["支付月份"]))
        if group == "未归属":
            unassigned_rows += 1
        if str(row["是否退款"]) == "1":
            if converted_user:
                refunded_order_users.add(converted_user)
        if row["是否激励"] == "是":
            if converted_user:
                incentive_order_users.add(converted_user)

    return {
        "groupStats": group_stats,
        "groupConvertedUsers": group_converted_users,
        "groupContributorStats": group_contributor_stats,
        "groupMeta": group_meta,
        "contributorRankStats": contributor_rank_stats,
        "contributorModules": contributor_modules,
        "contributorGroups": contributor_groups,
        "dailyStats": daily_stats,
        "subjectStats": subject_stats,
        "activityStats": activity_stats,
        "total": total,
        "months": months,
        "repairedRows": repaired_rows,
        "unassignedRows": unassigned_rows,
        "refundedOrders": len(refunded_order_users),
        "incentiveOrders": len(incentive_order_users),
        "filteredRows": filtered_rows,
        "activityFilteredRows": activity_filtered_rows,
        "convertedUsers": len(converted_users),
    }


def parse_arenas(group_stats, group_converted_users, group_contributor_stats, group_meta):
    wb = load_workbook(GROUPING_PATH, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    arenas = []
    arena_index = 0
    members_seen = set()
    missing_members = set()

    for source_row, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        data = dict(zip(headers, values))
        if not data.get("组号"):
            continue
        module = clean_name(data.get("模块"), "未知模块")
        if module not in ALLOWED_MODULES:
            continue

        members = []
        for index in (1, 2, 3):
            name = data.get(f"成员{index}_小组长")
            if not name:
                continue
            name = clean_name(name)
            role = clean_name(data.get(f"成员{index}_角色"), "参赛方")
            stats = round_stat(group_stats[name])
            meta = group_meta[name]
            n2_manager = meta["n2"].most_common(1)[0][0] if meta["n2"] else clean_name(data.get("n2"), "未知N2")
            manager = meta["manager"].most_common(1)[0][0] if meta["manager"] else n2_manager
            battle_power = round(stats["gross"] / 10000, 1)
            contributors = [
                {"name": contributor_name, **round_stat(contributor_stats)}
                for contributor_name, contributor_stats in group_contributor_stats[name].items()
                if contributor_stats["_orderUsers"]
            ]
            contributors.sort(key=lambda item: (item["net"], item["orders"]), reverse=True)
            members.append({
                "slot": index,
                "role": role,
                "name": name,
                "businessLine": clean_name(data.get(f"成员{index}_业务线"), "未知业务线"),
                "n2Manager": n2_manager,
                "manager": manager,
                "span": data.get(f"成员{index}_管理幅度") or 0,
                "teamSize": data.get(f"成员{index}_团队人数") or 0,
                "battlePower": battle_power,
                "convertedUsers": len(group_converted_users[name]),
                "contributors": contributors,
                **stats,
            })
            members_seen.add(name)
            if stats["orders"] == 0:
                missing_members.add(name)

        if not members:
            continue
        arena_index += 1

        sorted_members = sorted(members, key=lambda item: (item["gross"], item["net"], item["orders"]), reverse=True)
        winner_count = 2 if len(sorted_members) >= 3 else 1
        is_active = sum(item["orders"] for item in members) > 0
        winning_members = sorted_members[:winner_count] if is_active else []
        winning_names = {item["name"] for item in winning_members}
        for rank, member in enumerate(sorted_members, start=1):
            member["gmvRank"] = rank
            member["isWinner"] = member["name"] in winning_names
        winner = sorted_members[0]
        runner_up = sorted_members[1] if len(sorted_members) > 1 else None
        cutoff = sorted_members[winner_count - 1]
        next_member = sorted_members[winner_count] if len(sorted_members) > winner_count else None
        margin = cutoff["gross"] - (next_member["gross"] if next_member else 0) if is_active else 0
        defender = next((item for item in members if item["role"] == "守擂方"), members[0])
        challengers = [item for item in members if item["name"] != defender["name"]]
        challenger_stat = {
            "role": "攻擂阵营",
            "name": " / ".join(item["name"] for item in challengers) if challengers else "暂无攻擂方",
            "n2Manager": " / ".join(dict.fromkeys(item["n2Manager"] for item in challengers)) if challengers else "未知N2",
            "orders": sum(item["orders"] for item in challengers),
            "gross": round(sum(item["gross"] for item in challengers), 2),
            "refund": round(sum(item["refund"] for item in challengers), 2),
            "net": round(sum(item["net"] for item in challengers), 2),
        }
        challenger_stat["battlePower"] = round(challenger_stat["gross"] / 10000, 1)
        defender_won = defender["name"] in winning_names
        is_three_way = len(members) >= 3

        arenas.append({
            "id": f"A{arena_index:03d}",
            "sourceRow": source_row,
            "module": module,
            "commander": clean_name(data.get("n2"), "未知N2"),
            "groupNo": int(data["组号"]),
            "pkType": clean_name(data.get("PK类型"), f"{len(members)}人PK"),
            "matchName": clean_name(data.get("PK名单"), "未命名擂台"),
            "members": members,
            "winner": winner,
            "runnerUp": runner_up,
            "winningMembers": winning_members,
            "winnerCount": winner_count,
            "isThreeWay": is_three_way,
            "isActive": is_active,
            "cutoffGross": round(cutoff["gross"], 2),
            "totalNet": round(sum(item["net"] for item in members), 2),
            "totalGross": round(sum(item["gross"] for item in members), 2),
            "totalOrders": sum(item["orders"] for item in members),
            "margin": round(margin, 2),
            "defender": defender,
            "challengerSide": challenger_stat,
            "defenderWon": defender_won,
            "resultLabel": "待开战" if not is_active else ("GMV前二晋级" if is_three_way else ("守擂成功" if defender_won else "攻擂成功")),
            "intensity": round(sum(item["gross"] for item in members) / 10000, 1),
        })

    return arenas, members_seen, missing_members


def rank_items(items, key, limit):
    return sorted(items, key=key, reverse=True)[:limit]


def build_report():
    csv_data = parse_csv_stats()
    arenas, members_seen, missing_members = parse_arenas(
        csv_data["groupStats"],
        csv_data["groupConvertedUsers"],
        csv_data["groupContributorStats"],
        csv_data["groupMeta"],
    )

    member_rank = []
    for name in members_seen:
        stats = round_stat(csv_data["groupStats"][name])
        member_rank.append({
            "name": name,
            "battlePower": round(stats["gross"] / 10000, 1),
            "convertedUsers": len(csv_data["groupConvertedUsers"][name]),
            **stats,
        })
    member_rank.sort(key=lambda item: (-item["gross"], -item["net"], -item["orders"], item["name"]))

    member_modules = {}
    for arena in arenas:
        for member in arena["members"]:
            member_modules[member["name"]] = arena["module"]

    module_champions = []
    for module_name in ALLOWED_MODULE_ORDER:
        candidates = [item for item in member_rank if member_modules.get(item["name"]) == module_name]
        if candidates:
            module_champions.append({"module": module_name, **candidates[0]})

    pending_teams = []
    for module_name in ALLOWED_MODULE_ORDER:
        names = sorted(
            {
                member["name"]
                for arena in arenas
                if arena["module"] == module_name
                for member in arena["members"]
                if member["orders"] == 0
            }
        )
        pending_teams.append({
            "module": module_name,
            "count": len(names),
            "names": names,
        })

    module_map = defaultdict(lambda: {"arenas": 0, "defenderWins": 0, "challengerWins": 0, "gross": 0.0, "net": 0.0, "orders": 0})
    commander_map = defaultdict(lambda: {"arenas": 0, "wins": 0, "gross": 0.0, "net": 0.0, "orders": 0})
    for arena in arenas:
        module = module_map[arena["module"]]
        module["arenas"] += 1
        module["gross"] += arena["totalGross"]
        module["net"] += arena["totalNet"]
        module["orders"] += arena["totalOrders"]
        commander = commander_map[arena["commander"]]
        commander["arenas"] += 1
        commander["gross"] += arena["totalGross"]
        commander["net"] += arena["totalNet"]
        commander["orders"] += arena["totalOrders"]

        if arena["isActive"]:
            if arena["defenderWon"]:
                module["defenderWins"] += 1
            else:
                module["challengerWins"] += 1
            commander["wins"] += 1

    module_rank = [
        {
            "name": name,
            "arenas": value["arenas"],
            "defenderWins": value["defenderWins"],
            "challengerWins": value["challengerWins"],
            "gross": round(value["gross"], 2),
            "net": round(value["net"], 2),
            "orders": value["orders"],
        }
        for name, value in module_map.items()
    ]
    module_rank.sort(key=lambda item: item["gross"], reverse=True)

    commander_rank = [
        {
            "name": name,
            "arenas": value["arenas"],
            "wins": value["wins"],
            "gross": round(value["gross"], 2),
            "net": round(value["net"], 2),
            "orders": value["orders"],
        }
        for name, value in commander_map.items()
    ]
    commander_rank.sort(key=lambda item: item["gross"], reverse=True)

    contributor_rank = []
    for name, stat in csv_data["contributorRankStats"].items():
        rounded = round_stat(stat)
        if not rounded["orders"]:
            continue
        module = csv_data["contributorModules"][name].most_common(1)[0][0] if csv_data["contributorModules"][name] else "未知模块"
        group = csv_data["contributorGroups"][name].most_common(1)[0][0] if csv_data["contributorGroups"][name] else "未归属"
        contributor_rank.append({
            "name": name,
            "module": module,
            "group": group,
            "battlePower": round(rounded["gross"] / 10000, 1),
            **rounded,
        })
    contributor_rank.sort(key=lambda item: (-item["gross"], -item["net"], -item["orders"], item["name"]))

    daily = []
    cumulative = 0.0
    for date, stat in sorted(csv_data["dailyStats"].items()):
        rounded = round_stat(stat)
        cumulative += rounded["net"]
        daily.append({"date": date, **rounded, "cumulativeNet": round(cumulative, 2)})

    subjects = [
        {"name": name, **round_stat(stat)}
        for name, stat in csv_data["subjectStats"].items()
    ]
    subjects.sort(key=lambda item: item["net"], reverse=True)

    total = round_stat(csv_data["total"])
    activity_channels = []
    for activity_id in sorted(ALLOWED_ACTIVITY_IDS, key=int):
        stat = round_stat(csv_data["activityStats"][activity_id])
        activity_channels.append({
            "id": activity_id,
            "name": ACTIVITY_NAMES[activity_id],
            **stat,
        })
    activity_channels.sort(key=lambda item: (-item["gross"], -item["orders"], int(item["id"])))

    defender_wins = sum(1 for arena in arenas if arena["isActive"] and arena["defenderWon"])
    challenger_wins = sum(1 for arena in arenas if arena["isActive"] and not arena["defenderWon"])
    active_arenas = sum(1 for arena in arenas if arena["totalOrders"] > 0)

    report = {
        "meta": {
            "sourceFile": DATA_PATH.name,
            "groupingFile": GROUPING_PATH.name,
            "month": sorted(csv_data["months"])[0] if csv_data["months"] else "未知月份",
            "dateRange": {
                "start": daily[0]["date"] if daily else "",
                "end": daily[-1]["date"] if daily else "",
            },
            "generatedAt": datetime.now(timezone.utc).isoformat(),
            "repairedRows": csv_data["repairedRows"],
            "filteredRows": csv_data["filteredRows"],
            "activityFilteredRows": csv_data["activityFilteredRows"],
            "allowedActivityIds": sorted(ALLOWED_ACTIVITY_IDS, key=int),
        },
        "summary": {
            **total,
            "arenas": len(arenas),
            "activeArenas": active_arenas,
            "pkMembers": len(members_seen),
            "matchedMembers": len(members_seen - missing_members),
            "missingMembers": len(missing_members),
            "defenderWins": defender_wins,
            "challengerWins": challenger_wins,
            "threeWayArenas": sum(1 for arena in arenas if arena["isThreeWay"]),
            "unassignedRows": csv_data["unassignedRows"],
            "refundedOrders": csv_data["refundedOrders"],
            "incentiveOrders": csv_data["incentiveOrders"],
            "incentiveRate": round(csv_data["incentiveOrders"] / total["orders"], 4) if total["orders"] else 0,
            "convertedUsers": csv_data["convertedUsers"],
        },
        "champion": member_rank[0] if member_rank else None,
        "moduleChampions": module_champions,
        "pendingTeams": pending_teams,
        "arenas": arenas,
        "topArenas": rank_items(arenas, lambda item: item["totalGross"], 10),
        "closeArenas": sorted([arena for arena in arenas if arena["totalOrders"] > 0], key=lambda item: item["margin"])[:8],
        "memberRank": member_rank[:30],
        "contributorRank": contributor_rank,
        "moduleRank": module_rank,
        "commanderRank": commander_rank[:12],
        "daily": daily,
        "subjects": subjects[:8],
        "activityChannels": activity_channels,
        "missingMembers": sorted(missing_members),
    }
    return report


report = build_report()
OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
report_json = json.dumps(report, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
report_payload = base64.b64encode(gzip.compress(report_json, compresslevel=9, mtime=0)).decode("ascii")
OUTPUT_PATH.write_text(
    """window.PK_ARENA_DATA_READY = (async () => {
  const binary = atob("%s");
  const bytes = Uint8Array.from(binary, (character) => character.charCodeAt(0));
  const stream = new Blob([bytes]).stream().pipeThrough(new DecompressionStream("gzip"));
  window.PK_ARENA_DATA = JSON.parse(await new Response(stream).text());
  return window.PK_ARENA_DATA;
})();
"""
    % report_payload,
    encoding="utf-8",
)
print(f"Wrote {OUTPUT_PATH}")
print(f"{report['summary']['arenas']} arenas, champion {report['champion']['name']}, matched {report['summary']['matchedMembers']}/{report['summary']['pkMembers']}, three-way {report['summary']['threeWayArenas']}")
