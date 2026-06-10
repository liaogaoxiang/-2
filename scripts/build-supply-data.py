import json
from collections import defaultdict
from datetime import datetime
from pathlib import Path
import sys

try:
    from openpyxl import load_workbook
except ImportError as exc:
    raise SystemExit("openpyxl is required. Use the bundled Codex Python runtime.") from exc


DATA_PATH = Path(sys.argv[1] if len(sys.argv) > 1 else "/Users/gaoxiangliao/Desktop/测试数据.xlsx")
ARCH_PATH = Path(sys.argv[2] if len(sys.argv) > 2 else "/Users/gaoxiangliao/Desktop/精品架构映射.xlsx")
MATCH_PATH = Path(sys.argv[3] if len(sys.argv) > 3 else "/Users/gaoxiangliao/Desktop/队伍匹配.xlsx")
OUTPUT_PATH = Path("data/supply-data.js")

OFFICIAL_VALID_FROM = datetime(2026, 6, 11, 0, 0, 0)
VALID_FROM = OFFICIAL_VALID_FROM
VALID_TO = datetime(2026, 6, 25, 23, 59, 59)

ALLOWED_ACTIVITY_IDS = {
    "10216", "10227", "10240", "10275", "10390", "10404", "10418", "10435",
    "10496", "10528", "10553", "10601", "10628", "10717", "10719", "10380",
    "10294", "10777", "10819", "10849", "10830", "10896", "10890", "10928",
}

TIERS = [
    {"threshold": 20, "cardValue": 200, "item": "急速草鞋", "rarity": "疾行"},
    {"threshold": 30, "cardValue": 400, "item": "攻速手套", "rarity": "精良"},
    {"threshold": 40, "cardValue": 700, "item": "破茧斗篷", "rarity": "史诗"},
    {"threshold": 50, "cardValue": 1000, "item": "辉月法杖", "rarity": "传说"},
    {"threshold": 70, "cardValue": 1500, "item": "无尽战刃", "rarity": "神装"},
]


def clean(value):
    text = str(value or "").strip()
    if text.lower() in {"", "-", "null", "none", "nan", "其它", "其他"}:
        return ""
    return text


def amount(value):
    if value in (None, ""):
        return 0.0
    try:
        return float(str(value).replace(",", ""))
    except ValueError:
        return 0.0


def as_datetime(value):
    if isinstance(value, datetime):
        return value
    text = str(value or "").strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(text[: len(fmt)], fmt)
        except ValueError:
            continue
    return None


def load_match_leaders(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    index = {name: pos for pos, name in enumerate(headers)}
    leaders = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if clean(row[index["模块"]]) != "学习规划部":
            continue
        n2 = clean(row[index["n2"]]) or "未归属"
        group_no = clean(row[index["组号"]])
        for slot in (1, 2, 3):
            name = clean(row[index[f"成员{slot}_小组长"]])
            if not name:
                continue
            leaders[name] = {
                "name": name,
                "manager": n2,
                "groupNo": group_no,
                "role": clean(row[index[f"成员{slot}_角色"]]) or "参赛方",
                "pkType": clean(row[index["PK类型"]]),
                "declaredTeamSize": int(row[index[f"成员{slot}_团队人数"]] or 0),
            }

    return leaders


def load_planning_architecture(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    index = {name: pos for pos, name in enumerate(headers)}
    team_sizes = defaultdict(int)
    planning_rows = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        department = " ".join(clean(row[index[name]]) for name in ("department_1", "department_2", "department_3"))
        if "学习规划部" not in department:
            continue
        planning_rows += 1
        leader = clean(row[index["n1"]])
        if not leader:
            continue
        team_sizes[leader] += 1

    return team_sizes, planning_rows


def load_order_hits(path, leaders):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    index = {name: pos for pos, name in enumerate(headers)}
    hits = defaultdict(lambda: {"partners": set(), "orders": set(), "gross": 0.0, "net": 0.0})
    stats = {
        "sourceRows": 0,
        "validRows": 0,
        "beforeWindowRows": 0,
        "afterWindowRows": 0,
        "activityFilteredRows": 0,
        "invalidRows": 0,
        "matchedRows": 0,
    }
    max_payment = None
    leader_names = set(leaders)

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(value not in (None, "") for value in row):
            continue
        stats["sourceRows"] += 1
        activity_id = clean(row[index["活动ID"]])
        if activity_id not in ALLOWED_ACTIVITY_IDS:
            stats["activityFilteredRows"] += 1
            continue
        paid_at = as_datetime(row[index["支付时间"]])
        if paid_at:
            max_payment = paid_at if max_payment is None or paid_at > max_payment else max_payment
        if not paid_at or paid_at < VALID_FROM:
            stats["beforeWindowRows"] += 1
            continue
        if paid_at > VALID_TO:
            stats["afterWindowRows"] += 1
            continue
        if clean(row[index["是否违规"]]) in {"是", "1", "true", "True"}:
            stats["invalidRows"] += 1
            continue
        if str(row[index["是否退款"]] or "").strip() == "1":
            stats["invalidRows"] += 1
            continue
        gross = amount(row[index["收款金额"]])
        net = amount(row[index["净收金额"]])
        if gross <= 0 or net <= 0:
            stats["invalidRows"] += 1
            continue

        stats["validRows"] += 1
        partner = clean(row[index["业绩归属人姓名"]])
        leader = clean(row[index["小组长_来自人员架构表"]])
        if leader not in leader_names or not partner:
            continue
        order_user = clean(row[index["被邀请人ID"]]) or clean(row[index["订单编号"]])
        hits[leader]["partners"].add(partner)
        if order_user:
            hits[leader]["orders"].add(order_user)
        hits[leader]["gross"] += gross
        hits[leader]["net"] += net
        stats["matchedRows"] += 1

    return hits, stats, max_payment


def reached_tier(coverage):
    current = None
    for tier in TIERS:
        if coverage >= tier["threshold"]:
            current = tier
    return current


def next_tier(coverage):
    for tier in TIERS:
        if coverage < tier["threshold"]:
            return tier
    return None


def build_payload():
    match_leaders = load_match_leaders(MATCH_PATH)
    team_sizes, architecture_rows = load_planning_architecture(ARCH_PATH)
    hits, order_stats, max_payment = load_order_hits(DATA_PATH, match_leaders)
    leaders = []

    for leader, meta in match_leaders.items():
        hit = hits[leader]
        denominator = team_sizes.get(leader, 0)
        covered = len(hit["partners"])
        coverage = round(covered / denominator * 100, 1) if denominator else 0
        current = reached_tier(coverage)
        upcoming = next_tier(coverage)
        leaders.append({
            "name": leader,
            "manager": meta["manager"],
            "groupNo": meta["groupNo"],
            "role": meta["role"],
            "pkType": meta["pkType"],
            "declaredTeamSize": meta["declaredTeamSize"],
            "teamSize": denominator,
            "coveredPartners": covered,
            "coverage": coverage,
            "orders": len(hit["orders"]),
            "gross": round(hit["gross"], 2),
            "net": round(hit["net"], 2),
            "hitPartners": sorted(hit["partners"]),
            "currentTier": current,
            "nextTier": upcoming,
            "partnersNeeded": max(0, int((upcoming["threshold"] * denominator + 99) // 100) - covered) if upcoming else 0,
        })

    leaders.sort(key=lambda item: (-item["coverage"], -item["coveredPartners"], -item["gross"], item["name"]))
    active_leaders = sum(1 for item in leaders if item["coveredPartners"] > 0)
    summary = {
        "leaders": len(leaders),
        "partners": architecture_rows,
        "displayedLeaderPartners": sum(item["teamSize"] for item in leaders),
        "coveredPartners": sum(item["coveredPartners"] for item in leaders),
        "activeLeaders": active_leaders,
        "averageCoverage": round(sum(item["coverage"] for item in leaders) / len(leaders), 1) if leaders else 0,
        "maxCoverage": leaders[0]["coverage"] if leaders else 0,
        "totalOrders": sum(item["orders"] for item in leaders),
        "totalGross": round(sum(item["gross"] for item in leaders), 2),
        "missingArchitectureLeaders": sum(1 for item in leaders if item["teamSize"] == 0),
        **order_stats,
    }

    return {
        "meta": {
            "title": "荣耀擂台赛补给包",
            "slogan": "装备升上去，战斗赢下来！",
            "activityTime": "6月1日-6月25日",
            "officialValidRule": "仅统计支付时间为2026-06-11至2026-06-25的有效出单伙伴",
            "currentStatRule": "正式口径: 当前页面按2026-06-11至2026-06-25统计",
            "sourceFile": DATA_PATH.name,
            "architectureFile": ARCH_PATH.name,
            "matchFile": MATCH_PATH.name,
            "statStart": VALID_FROM.strftime("%Y-%m-%d"),
            "officialStart": OFFICIAL_VALID_FROM.strftime("%Y-%m-%d"),
            "statEnd": VALID_TO.strftime("%Y-%m-%d"),
            "latestPaymentTime": max_payment.strftime("%Y-%m-%d %H:%M:%S") if max_payment else "",
            "generatedAt": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        },
        "tiers": TIERS,
        "summary": summary,
        "leaders": leaders,
    }


payload = build_payload()
OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
OUTPUT_PATH.write_text(
    "window.PK_SUPPLY_DATA = " + json.dumps(payload, ensure_ascii=False, separators=(",", ":")) + ";\n",
    encoding="utf-8",
)

print(f"Wrote {OUTPUT_PATH}")
print(
    f"{payload['summary']['leaders']} leaders, "
    f"{payload['summary']['coveredPartners']}/{payload['summary']['partners']} covered, "
    f"valid rows {payload['summary']['validRows']}, matched rows {payload['summary']['matchedRows']}"
)
